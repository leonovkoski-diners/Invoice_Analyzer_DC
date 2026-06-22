"""
Invoice Analyzer — Helix-K Excel Export Engine
================================================
Writes a verified InvoiceRecord to a .xlsx file formatted exactly
to the Zonel Helix-K import specification.

Column order, header names, date format (DD/MM/YYYY), and decimal
separator are hardcoded to the Helix-K spec. Do not change these
without updating the Helix-K import template configuration.
"""

from __future__ import annotations

import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

from pipeline.schema import InvoiceRecord

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helix-K column specification — ORDER MATTERS
# ---------------------------------------------------------------------------

HELIX_COLUMNS = [
    "Supplier Name",
    "VAT ID",
    "Invoice Ref",
    "Invoice Date",    # DD/MM/YYYY
    "Due Date",        # DD/MM/YYYY
    "Currency",
    "Net Amount",      # 2dp decimal
    "VAT Rate (%)",    # 2dp decimal
    "VAT Amount",      # 2dp decimal
    "Gross Amount",    # 2dp decimal
    "Payment Reference",
    "_Validation Flags",  # Hidden audit column — always last
]


def _fmt_date(iso_date: Optional[str]) -> str:
    """Convert ISO 8601 date to Helix-K format (DD/MM/YYYY)."""
    if not iso_date:
        return ""
    try:
        d = datetime.strptime(iso_date, "%Y-%m-%d")
        return d.strftime("%d/%m/%Y")
    except ValueError:
        return iso_date  # Return as-is if not normalized


def _fmt_decimal(v: Optional[Decimal], dp: int = 2) -> str:
    """Format a Decimal to a string with fixed decimal places."""
    if v is None:
        return ""
    return f"{v:.{dp}f}"


def export_to_helix(
    record: InvoiceRecord,
    validation_flags: list[str],
    output_dir: str | Path,
    output_filename: Optional[str] = None,
) -> Path:
    """
    Write a verified InvoiceRecord to a Helix-K compatible .xlsx file.

    Returns the path to the generated file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        ) from e

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate filename: YYYY-MM-DD_VendorName_INV-XXXX.xlsx
    if not output_filename:
        safe_vendor = "".join(
            c if c.isalnum() or c in "-_ " else "_"
            for c in record.vendor_name
        )[:30].strip()
        safe_invoice = "".join(
            c if c.isalnum() or c in "-_" else "_"
            for c in record.invoice_number
        )[:20]
        date_prefix = (record.invoice_date or "").replace("-", "")[:8]
        output_filename = f"{date_prefix}_{safe_vendor}_{safe_invoice}.xlsx"

    output_path = output_dir / output_filename

    # ── Build workbook ───────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Helix-K Import"

    # ── Header row styling ───────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A1A6E", end_color="1A1A6E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(HELIX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 30

    # ── One row per line item ────────────────────────────────────────────
    # Helix-K expects one row per line item, with header fields repeated
    flag_str = "; ".join(validation_flags) if validation_flags else "CLEAN"

    rows_written = 0
    for line in record.line_items:
        row = rows_written + 2  # Row 1 is header

        row_data = [
            record.vendor_name,
            record.vendor_vat_id or "",
            record.invoice_number,
            _fmt_date(record.invoice_date),
            _fmt_date(record.due_date),
            record.currency,
            _fmt_decimal(record.subtotal),
            _fmt_decimal(record.tax_rate),
            _fmt_decimal(record.tax_amount),
            _fmt_decimal(record.total),
            record.payment_reference or "",
            flag_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=False)

            # Highlight validation flag column if there are issues
            if col_idx == len(HELIX_COLUMNS) and validation_flags:
                cell.fill = PatternFill(
                    start_color="FEF3E2", end_color="FEF3E2", fill_type="solid"
                )
                cell.font = Font(color="7A4100", size=9)

        rows_written += 1

    # ── Column widths ────────────────────────────────────────────────────
    col_widths = {
        "Supplier Name": 30,
        "VAT ID": 16,
        "Invoice Ref": 18,
        "Invoice Date": 14,
        "Due Date": 14,
        "Currency": 10,
        "Net Amount": 14,
        "VAT Rate (%)": 12,
        "VAT Amount": 14,
        "Gross Amount": 14,
        "Payment Reference": 22,
        "_Validation Flags": 50,
    }

    for col_idx, col_name in enumerate(HELIX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 16)

    # Hide the validation flags column (it's an audit trail, not for import)
    # Comment this out if you want it visible
    # ws.column_dimensions[get_column_letter(len(HELIX_COLUMNS))].hidden = True

    # ── Freeze header row ────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(str(output_path))
    logger.info(f"Exported to: {output_path}")
    return output_path
